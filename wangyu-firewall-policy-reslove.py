from openpyxl import Workbook
import sys
from re import findall,sub


def readFile(filename):
    i = 1
    wb = Workbook()
    ws = wb.active
    ws.tile = "policy"
    ws['A1'] = "策略id"
    ws['B1'] = "策略名"
    ws['C1'] = "源地址"
    ws['D1'] = "目的地址"
    ws['E1'] = "服务"
    ws['F1'] = "动作"
    ws['G1'] = "策略状态"
    ws['H1'] = '备注'
    f = open(filename, 'r',)
    file_list = f.readlines()
    policy_list = [x for x in file_list if "setdb rule policyinfo" in x]
    for policy in policy_list:
        policy = policy.replace("setdb rule policyinfo ", "")
        policy_string = findall(r'comment\s\"(.*?)\"', policy)[0].strip()
        policy = sub(r'comment\s\"(.*?)\"', "comment "+policy_string, policy)
        new_policy_list = policy.split(" ")
        new_policy_dict = dict(zip(new_policy_list[0::2], new_policy_list[1::2]))
        # 将列表转换为字典，参考https://www.cnblogs.com/cantin-python-notes/p/9243067.html
        if new_policy_dict["type"] == '"1"' or new_policy_dict["type"] == '"2"':
            i += 1
            if "id" in new_policy_dict.keys():
                ws['A%d' % i] = eval(new_policy_dict['id'])
            else:
                ws['A%d' % i] = "-"
            if "name" in new_policy_dict.keys():
                ws['B%d' % i] = eval(new_policy_dict['name'])
            else:
                ws['B%d' % i] = "-"
            if "saddrid" in new_policy_dict.keys():
                if "saddrtype" in new_policy_dict.keys():
                    if new_policy_dict["saddrtype"] == '"2"':
                        ws['C%d' % i] = getSingleIp(new_policy_dict["saddrid"], filename)
                    if new_policy_dict["saddrtype"] == '"7"':
                        ws['C%d' % i] = "any"
                    if new_policy_dict["saddrtype"] == '"3"':
                        ip_new = ""
                        for y in getGIp(new_policy_dict["saddrid"], filename):
                            ip_new = ip_new + "  " + getSingleIp(y, filename)
                        ws['C%d' % i] = ip_new
                    if new_policy_dict["saddrtype"] == '"9"':
                        ip_new = ""
                        for y in getTemIp(new_policy_dict["saddrid"], filename):
                            ip_new = ip_new + "  " + getSingleIp(y, filename)
                        ws['C%d' % i] = ip_new
                else:
                    ws['C%d' % i] = "-"
            if "daddrid" in new_policy_dict.keys():
                if "daddrtype" in new_policy_dict.keys():
                    if new_policy_dict["daddrtype"] == '"2"':
                        ws['D%d' % i] = getSingleIp(new_policy_dict["daddrid"], filename)
                    if new_policy_dict["daddrtype"] == '"7"':
                        ws['D%d' % i] = "any"
                    if new_policy_dict["daddrtype"] == '"3"':
                        ip_new = ""
                        for y in getGIp(new_policy_dict["daddrid"], filename):
                            ip_new = ip_new + "  " + getSingleIp(y, filename)
                        ws['D%d' % i] = ip_new
                    if new_policy_dict["daddrtype"] == '"9"':
                        ip_new = ""
                        for y in getTemIp(new_policy_dict["daddrid"], filename):
                            ip_new = ip_new + "  " + getSingleIp(y, filename)
                        ws['D%d' % i] = ip_new
            else:
                ws['D%d' % i] = "-"
            if "serviceid" in new_policy_dict.keys():
                if "servicetype" in new_policy_dict.keys():
                    if new_policy_dict["servicetype"] == '"6"':
                        ws['E%d' % i] = "any"
                    if new_policy_dict["servicetype"] == '"1"':
                        service_new = ""
                        for y in getAllService(new_policy_dict["serviceid"], filename):
                            service_new = service_new + "  " + y.replace('"', "")
                        ws['E%d' % i] = service_new
                    if new_policy_dict["servicetype"] == '"4"':
                        service_new = ""
                        for y in getOneService(new_policy_dict["serviceid"], filename):
                            service_new = service_new + "  " + y.replace('"', "")
                        for z in getOneRuleService(new_policy_dict["serviceid"], filename):
                            service_new = service_new + "  " + z.replace('"', "")
                        ws['E%d' % i] = service_new
                    if new_policy_dict["servicetype"] == '"5"':
                        service_new = ""
                        for y in getServicePort(new_policy_dict["serviceid"], filename):
                            for z in getAllService(y, filename):
                                service_new = service_new + "  " + z.replace('"', "")
                        ws['E%d' % i] = service_new
            else:
                ws['E%d' % i] = "-"
            if "active" in new_policy_dict.keys():
                if new_policy_dict["type"] == '"2"': 
                    if new_policy_dict['active'] == '"1"':
                        ws['F%d' % i] = "禁止"
                    if new_policy_dict['active'] == '"0"':
                        ws['F%d' % i] = "允许"
                if new_policy_dict["type"] == '"1"':
                    if new_policy_dict['POLICY_W'] != '"0"':
                        if new_policy_dict['active'] == '"1"':
                            ws['F%d' % i] = "允许"
                        if new_policy_dict['active'] == '"0"':
                            ws['F%d' % i] = "禁止"
                    if new_policy_dict['POLICY_W'] == '"0"':
                        if new_policy_dict['active'] == '"1"':
                            ws['F%d' % i] = "禁止"
                        if new_policy_dict['active'] == '"0"':
                            ws['F%d' % i] = "允许"
            else:
                ws['F%d' % i] = "-"
            if "POLICY_W" in new_policy_dict.keys():
                if new_policy_dict['POLICY_W'] != '"0"':
                    ws['G%d' % i] = "启用"
                if new_policy_dict['POLICY_W'] == '"0"':
                    ws['G%d' % i] = "未启用"
            else:
                ws['G%d' % i] = "-"
            if "comment" in new_policy_dict.keys():
                ws['H%d' % i] = new_policy_dict['comment'].replace('"', "")
            else:
                ws['H%d' % i] = "-"
    wb.save('下联防火墙-10.1.51.84-20201029.xlsx')


def getAddress(filename):
    final_address_list = []
    f = open(filename, 'r', encoding='utf-8')
    ip_file_list = f.readlines()
    ip_range_list = [x for x in ip_file_list if "setdb rule address" in x]
    for ip_list in ip_range_list:
        ip_list = ip_list.replace("setdb rule address ", "")
        new_ip_list = ip_list.split(" ")
        new_ip_dict = dict(zip(new_ip_list[0::2], new_ip_list[1::2]))
        if 'type' in new_ip_dict.keys() and new_ip_dict['type'] == '"1"':
            new_ip_dict["ip_new"] = new_ip_dict["ip"] + "/" + new_ip_dict["mask"]
        if 'type' in new_ip_dict.keys() and new_ip_dict['type'] == '"2"':
            new_ip_dict["ip_new"] = new_ip_dict["ip"] + "-" + new_ip_dict["mask"]
        final_address_list.append(new_ip_dict)
    return final_address_list


def getGroupId(filename):
    final_group_list = []
    f = open(filename, 'r', encoding='utf-8')
    group_id_file = f.readlines()
    group_id_list = [x for x in group_id_file if "setdb rule addrgrp" in x]
    for group_id in group_id_list:
        group_id = group_id.replace("setdb rule addrgrp ", "")
        new_group_id = group_id.split(" ")
        new_group_id_dict = dict(zip(new_group_id[0::2], new_group_id[1::2]))
        final_group_list.append(new_group_id_dict)
    return final_group_list


def getAddrmap(filename):
    final_addrmap_list = []
    f = open(filename, 'r', encoding='utf-8')
    addrmap_file_list = f.readlines()
    addrmap_id_list = [x for x in addrmap_file_list if "setdb rule addrmap" in x]
    for addrmap_id in addrmap_id_list:
        addrmap_id = addrmap_id.replace("setdb rule addrmap ","")
        new_addrmap_id = addrmap_id.split(" ")
        new_addrmap_id_dict = dict(zip(new_addrmap_id[0::2], new_addrmap_id[1::2]))
        final_addrmap_list.append(new_addrmap_id_dict)
    return final_addrmap_list


def getTemgroupId(filename):
    final_temgroup_list = []
    f = open(filename, 'r', encoding='utf-8')
    temgroup_id_file = f.readlines()
    temgroup_id_list = [x for x in temgroup_id_file if "setdb rule tempaddrgrp" in x]
    for temgroup_id in temgroup_id_list:
        temgroup_id = temgroup_id.replace("setdb rule tempaddrgrp ", "")
        new_temgroup_id = temgroup_id.split(" ")
        new_temgroup_id_dict = dict(zip(new_temgroup_id[0::2], new_temgroup_id[1::2]))
        final_temgroup_list.append(new_temgroup_id_dict)
    return final_temgroup_list


def getTemMap(filename):
    final_temmap_list = []
    f = open(filename, 'r', encoding='utf-8')
    temmap_file_list = f.readlines()
    temmap_id_list = [x for x in temmap_file_list if "setdb rule tempaddrmap" in x]
    for temmap_id in temmap_id_list:
        temmap_id = temmap_id.replace("setdb rule tempaddrmap ", "")
        new_temmap_id = temmap_id.split(" ")
        new_temmap_id_dict = dict(zip(new_temmap_id[0::2], new_temmap_id[1::2]))
        final_temmap_list.append(new_temmap_id_dict)
    return final_temmap_list


def getService(filename):
    f = open(filename, 'r', encoding='utf-8')
    ip_file_list = f.readlines()
    final_service_list = []
    firewall_service_list = [x for x in ip_file_list if "setdb rule defaultservice" in x]
    for service_list in firewall_service_list:
        service_list = service_list.replace("setdb rule defaultservice ", "")
        new_service_list = service_list.split(" ")
        new_service_dict = dict(zip(new_service_list[0::2], new_service_list[1::2]))
        if new_service_dict["port"] == '"-1"':
            new_service_dict["service_new"] = new_service_dict["name"]
        if new_service_dict["port"] != '"-1"':
            new_service_dict["service_new"] = new_service_dict["name"] + "  " + new_service_dict["port"]
        final_service_list.append(new_service_dict)
    return final_service_list


def getDservice(filename):
    f = open(filename, 'r', encoding='utf-8')
    ip_file_list = f.readlines()
    final_service_dynamic_list = []
    firewall_service_dynamic_list = [x for x in ip_file_list if "setdb rule dynamic_service" in x]
    for service_dynamic_list in firewall_service_dynamic_list:
        service_dynamic_list = service_dynamic_list.replace("setdb rule dynamic_service ", "")
        new_service_dynamic_list = service_dynamic_list.split(" ")
        new_service_dynamic_dict = dict(zip(new_service_dynamic_list[0::2], new_service_dynamic_list[1::2]))
        if new_service_dynamic_dict["port"] == '"-1"':
            new_service_dynamic_dict["service_new"] = new_service_dynamic_dict["name"]
        if new_service_dynamic_dict["port"] != '"-1"':
            new_service_dynamic_dict["service_new"] = new_service_dynamic_dict["name"] + "  " +new_service_dynamic_dict["port"]
        final_service_dynamic_list.append(new_service_dynamic_dict)
    return final_service_dynamic_list


def getRuleService(filename):
    f = open(filename, 'r', encoding='utf-8')
    ip_file_list = f.readlines()
    final_rule_service_list = []
    firewall_rule_service_list = [x for x in ip_file_list if "setdb rule service " in x]
    for rule_service_list in firewall_rule_service_list:
        rule_service_list = rule_service_list.replace("setdb rule service ", "")
        new_rule_service_list = rule_service_list.split(" ")
        new_rule_service_dict = dict(zip(new_rule_service_list[0::2], new_rule_service_list[1::2]))
        if 'name' in new_rule_service_dict.keys():
            new_rule_service_dict["rule_service_name"] = new_rule_service_dict['name']
        rule_service_new_source = ""
        rule_service_new_des = ""
        for i in range(8):
            i += 1
            if 'slport%d' %i in new_rule_service_dict.keys() and 'shport%d' %i in new_rule_service_dict.keys():
                if new_rule_service_dict["slport%d" % i] != '"0"' and new_rule_service_dict["shport%d" % i] != '"65535"':
                    rule_service_new_source = rule_service_new_source + " " + new_rule_service_dict["slport%d" % i] + "-" + new_rule_service_dict["shport%d" % i]
                new_rule_service_dict["rule_service_source"] = rule_service_new_source
            if 'dlport%d' %i in new_rule_service_dict.keys() and 'dhport%d' %i in new_rule_service_dict.keys():
                if new_rule_service_dict["dlport%d" % i] != '"0"' and new_rule_service_dict["dhport%d" % i] != '"65535"':
                    rule_service_new_des = rule_service_new_des + " " + new_rule_service_dict["dlport%d" % i] + "-" + new_rule_service_dict["dhport%d" % i]
                new_rule_service_dict["rule_service_des"] = rule_service_new_des
        final_rule_service_list.append(new_rule_service_dict)
    return final_rule_service_list


def getSingleIp(id, filename):
    id_list = getAddress(filename)
    for ip_dict in id_list:
        if id in ip_dict.values():
            return ip_dict['ip_new'].replace('"', "")


def getGIp(id, filename):
    group_list = getGroupId(filename)
    addrmapid_list = []
    for group_dict in group_list:
        if id in group_dict.values():
            for addrmapid in getAddrmap(filename):
                if "addrgrpid" in addrmapid.keys():
                    if id == addrmapid["addrgrpid"]:
                        addrmapid_list.append(addrmapid["addrid"])
    return addrmapid_list


def getTemIp(id, filename):
    group_list = getTemgroupId(filename)
    addrmapid_list = []
    for group_dict in group_list:
        if id in group_dict.values():
            for tempaddrid in getTemMap(filename):
                if "tempaddrid" in tempaddrid.keys():
                    if id == tempaddrid["tempaddrgrpid"]:
                        addrmapid_list.append(tempaddrid["tempaddrid"])
    return addrmapid_list


def getAllService(id, filename):
    service_new_list = []
    service_list = getService(filename) + getDservice(filename)
    for service_dict in service_list:
        if id == service_dict["id"]:
            service = service_dict["service_new"]
            service_new_list.append(service)
    return service_new_list


def getOneService(id, filename):
    service_new_list = []
    service_list = getService(filename)
    for service_dict in service_list:
        if id == service_dict["id"]:
            service = service_dict["service_new"]
            service_new_list.append(service)
    return service_new_list


def getOneRuleService(id, filename):
    service_new_list = []
    service_list = getRuleService(filename)
    for service_dict in service_list:
        if id == service_dict["id"]:
            service = service_dict["rule_service_name"] + "  " + service_dict["rule_service_source"] + "  " + service_dict["rule_service_des"]
            service_new_list.append(service)
    return service_new_list


def getServiceGroupId(filename):
    final_service_group_list = []
    f = open(filename, 'r', encoding='utf-8')
    group_id_file = f.readlines()
    group_id_list = [x for x in group_id_file if "setdb rule servicegrp " in x]
    for group_id in group_id_list:
        group_id = group_id.replace("setdb rule servicegrp ", "")
        new_group_id = group_id.split(" ")
        new_group_id_dict = dict(zip(new_group_id[0::2], new_group_id[1::2]))
        final_service_group_list.append(new_group_id_dict)
    return final_service_group_list


def getServicemap(filename):
    final_service_map_list = []
    f = open(filename, 'r', encoding='utf-8')
    service_map_file_list = f.readlines()
    service_map_id_list = [x for x in service_map_file_list if "setdb rule servicemap " in x]
    for service_map_id in service_map_id_list:
        service_map_id = service_map_id.replace("setdb rule servicemap ", "")
        new_service_map_id = service_map_id.split(" ")
        new_service_map_id_dict = dict(zip(new_service_map_id[0::2], new_service_map_id[1::2]))
        final_service_map_list.append(new_service_map_id_dict)
    return final_service_map_list


def getServicePort(id, filename):
    service_port_list = getServiceGroupId(filename)
    service_map_id_list = []
    for service_port_dict in service_port_list:
        if id in service_port_dict.values():
            for servicemapid in getServicemap(filename):
                if "servicegrpid" in servicemapid.keys():
                    if id == servicemapid["servicegrpid"]:
                        service_map_id_list.append(servicemapid["serviceid"])
    return service_map_id_list


def main():
    readFile(sys.argv[1])


if __name__ == "__main__":
    main()
