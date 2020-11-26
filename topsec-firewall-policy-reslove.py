# coding=utf-8

from openpyxl import Workbook
import sys
from re import findall


def readFile(filename):
    ip_data_fin = ipData(filename)
    i = 1
    wb = Workbook()
    ws = wb.active
    ws.tile = "policy"
    ws['A1'] = "策略组"
    ws['B1'] = "描述"
    ws['C1'] = "源区域"
    ws['D1'] = "源地址"
    ws['E1'] = "目的区域"
    ws['F1'] = "目的地址"
    ws['G1'] = "服务"
    ws['H1'] = '域名'
    ws['I1'] = '控制'
    ws['J1'] = '策略状态'
    f = open(filename, 'r', encoding='utf-8')
    file_list = f.readlines()
    policy_list = [x for x in file_list if "firewall policy" in x]
    for policy in policy_list:
        policy = policy.replace("firewall policy add ", "")
        new_policy_list = policy.split("'")
        new_policy_list_on = [new_policy_list[y].replace(" ", "") if y % 2 != 0 else new_policy_list[y]
                              for y in range(len(new_policy_list))]
        ori_string = ""
        for string in new_policy_list_on:
            ori_string += string
        new_policy_list_on_fin = ori_string.split(" ")
        new_policy_dict = dict(zip(new_policy_list_on_fin[0::2], new_policy_list_on_fin[1::2]))
        # 将列表转换为字典，参考https://www.cnblogs.com/cantin-python-notes/p/9243067.html
        i += 1
        if "group_name" in new_policy_dict.keys():
            ws['A%d' % i] = new_policy_dict['group_name'].strip("'")
        else:
            ws['A%d' % i] = "-"
        if "comment" in new_policy_dict.keys():
            ws['B%d' % i] = new_policy_dict['comment']
        else:
            ws['B%d' % i] = "-"
        if "srcarea" in new_policy_dict.keys():
            ws['C%d' % i] = new_policy_dict['srcarea']
        else:
            ws['C%d' % i] = "-"
        if "src" in new_policy_dict.keys():
            policy_data = findall(r'\'(.*?)\'', policy)
            ip_name_data_string = ""
            for new_policy_data in policy_data:
                new_policy_data = new_policy_data.split(" ")
                for ip_data in new_policy_data:
                    for ip_name_data in ip_data_fin:
                        if ip_data in ip_name_data.values():
                            if ip_data in new_policy_dict['src']:
                                ip_name_data_string += ip_name_data["ip_new"]
            ws['D%d' % i] = ip_name_data_string
        else:
            ws['D%d' % i] = "-"
        if "dstarea" in new_policy_dict.keys():
            ws['E%d' % i] = new_policy_dict['dstarea']
        else:
            ws['E%d' % i] = "-"
        if "dst" in new_policy_dict.keys():
            policy_data = findall(r'\'(.*?)\'', policy)
            ip_name_data_string = ""
            for new_policy_data in policy_data:
                new_policy_data = new_policy_data.split(" ")
                for ip_data in new_policy_data:
                    for ip_name_data in ip_data_fin:
                        if ip_data in ip_name_data.values():
                            if ip_data in new_policy_dict['dst']:
                                ip_name_data_string += ip_name_data["ip_new"]
            ws['F%d' % i] = ip_name_data_string
        else:
            ws['F%d' % i] = "-"
        if "service" in new_policy_dict.keys():
            ws['G%d' % i] = new_policy_dict['service']
        else:
            ws['G%d' % i] = "-"
        if "domain" in new_policy_dict.keys():
            ws['H%d' % i] = new_policy_dict['domain']
        else:
            ws['H%d' % i] = "-"
        if "action" in new_policy_dict.keys():
            ws['I%d' % i] = new_policy_dict['action']
        else:
            ws['I%d' % i] = "-"
        if "enable" in new_policy_dict.keys():
            ws['J%d' % i] = "禁用"
        else:
            ws['J%d' % i] = "启用"
    wb.save('firewall-policy.xlsx')


def ipData(filename):
    final_new_ip_list = []
    final_new_host_list = []
    final_new_subnet_list = []
    f = open(filename, 'r', encoding='utf-8')
    ip_file_list = f.readlines()
    ip_range_list = [x for x in ip_file_list if "define range add" in x]
    for ip_list in ip_range_list:
        ip_list = ip_list.replace("define range add ", "")
        new_ip_list = ip_list.split(" ")
        new_ip_dict = dict(zip(new_ip_list[0::2], new_ip_list[1::2]))
        new_ip_dict["ip_new"] = new_ip_dict["ip1"] + "-" + new_ip_dict["ip2"]
        final_new_ip_list.append(new_ip_dict)
    ip_host_list = [x for x in ip_file_list if "define host add" in x]
    for host_list in ip_host_list:
        host_list = host_list.replace("define host add ", "")
        new_host_list = host_list.split(" ")
        new_host_dict = dict(zip(new_host_list[0::2], new_host_list[1::2]))
        new_host_dict["ip_new"] = new_host_dict["ipaddr"].replace("'", "") + " "
        final_new_host_list.append(new_host_dict)
    ip_subnet_list = [x for x in ip_file_list if "define subnet add" in x]
    for subnet_list in ip_subnet_list:
        subnet_list = subnet_list.replace("define subnet add ", "")
        new_subnet_list = subnet_list.split(" ")
        new_subnet_dict = dict(zip(new_subnet_list[0::2], new_subnet_list[1::2]))
        new_subnet_dict["ip_new"] = new_subnet_dict["ipaddr"] + "/" + new_subnet_dict["mask"]
        final_new_subnet_list.append(new_subnet_dict)
    return final_new_ip_list + final_new_host_list + final_new_subnet_list


def main():
    readFile(sys.argv[1])


if __name__ == "__main__":
    main()
